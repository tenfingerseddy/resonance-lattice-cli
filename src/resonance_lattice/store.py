# SPDX-License-Identifier: BUSL-1.1
"""Source Store — multi-resolution content storage for materialisation.

Stores the actual content that gets materialised after resonance identifies
bright spots. Content is stored at multiple resolutions:
  - Summary text (for Omega_1-2 landscape materialisation)
  - Relation triples (for Omega_3 structural materialisation)
  - Full text (for Omega_4-5 evidence materialisation)
  - Metadata (timestamp, source URI, authority, etc.)

Backed by SQLite for simplicity and portability.
"""

from __future__ import annotations

import csv
import json
import sqlite3
from pathlib import Path
from typing import Any


class SourceContent:
    """Multi-resolution content for a single source.

    Relations and metadata use lazy JSON deserialization — the raw JSON
    strings are stored and only parsed on first access.  This avoids
    ``json.loads`` overhead during batch retrieval when callers only
    need ``summary`` or ``full_text``.
    """

    __slots__ = (
        "source_id", "summary", "full_text",
        "_relations", "_metadata", "_relations_json", "_metadata_json",
    )

    def __init__(
        self,
        source_id: str,
        summary: str = "",
        relations: list[tuple[str, str, str]] | None = None,
        full_text: str = "",
        metadata: dict[str, Any] | None = None,
    ) -> None:
        self.source_id = source_id
        self.summary = summary
        self.full_text = full_text
        self._relations = relations
        self._metadata = metadata
        self._relations_json = json.dumps(relations) if relations is not None else "[]"
        self._metadata_json = json.dumps(metadata) if metadata is not None else "{}"

    @property
    def relations(self) -> list[tuple[str, str, str]]:
        if self._relations is None:
            self._relations = [tuple(r) for r in json.loads(self._relations_json)]
        return self._relations

    @relations.setter
    def relations(self, value: list[tuple[str, str, str]]) -> None:
        self._relations = value
        self._relations_json = json.dumps(value)

    @property
    def metadata(self) -> dict[str, Any]:
        if self._metadata is None:
            self._metadata = json.loads(self._metadata_json)
        return self._metadata

    @metadata.setter
    def metadata(self, value: dict[str, Any]) -> None:
        self._metadata = value
        self._metadata_json = json.dumps(value)

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_id": self.source_id,
            "summary": self.summary,
            "relations": self.relations,
            "full_text": self.full_text,
            "metadata": self.metadata,
        }

    def __repr__(self) -> str:
        return f"SourceContent(source_id={self.source_id!r}, summary={self.summary[:50]!r}...)"

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> SourceContent:
        return cls(
            source_id=data["source_id"],
            summary=data.get("summary", ""),
            relations=[tuple(r) for r in data.get("relations", [])],
            full_text=data.get("full_text", ""),
            metadata=data.get("metadata", {}),
        )

    @classmethod
    def _from_row(
        cls, source_id: str, summary: str,
        relations_json: str, full_text: str, metadata_json: str,
    ) -> SourceContent:
        """Fast constructor from a SQLite row — defers JSON parsing."""
        obj = cls.__new__(cls)
        obj.source_id = source_id
        obj.summary = summary
        obj.full_text = full_text
        obj._relations = None
        obj._metadata = None
        obj._relations_json = relations_json
        obj._metadata_json = metadata_json
        return obj


class SourceStore:
    """SQLite-backed multi-resolution source content store.

    Includes a read-through LRU cache for frequently accessed sources
    and SQLite memory-mapping for zero-copy reads.
    """

    _CACHE_SIZE = 512

    def __init__(self, path: str | Path = ":memory:") -> None:
        """
        Args:
            path: Path to SQLite database file, or ":memory:" for in-memory.
        """
        self.path = str(path)
        self._conn = sqlite3.connect(self.path, check_same_thread=False)
        self._conn.execute("PRAGMA journal_mode=WAL")
        self._conn.execute("PRAGMA mmap_size=268435456")  # 256 MB mmap
        self._create_tables()
        self._cache: dict[str, SourceContent] = {}

    def _create_tables(self) -> None:
        self._conn.execute("""
            CREATE TABLE IF NOT EXISTS sources (
                source_id TEXT PRIMARY KEY,
                summary TEXT NOT NULL DEFAULT '',
                relations TEXT NOT NULL DEFAULT '[]',
                full_text TEXT NOT NULL DEFAULT '',
                metadata TEXT NOT NULL DEFAULT '{}'
            )
        """)
        self._conn.commit()

    @property
    def count(self) -> int:
        cursor = self._conn.execute("SELECT COUNT(*) FROM sources")
        return cursor.fetchone()[0]

    def store(self, content: SourceContent) -> None:
        """Store or update source content.

        Args:
            content: The multi-resolution content to store.
        """
        self._conn.execute(
            """INSERT OR REPLACE INTO sources (source_id, summary, relations, full_text, metadata)
               VALUES (?, ?, ?, ?, ?)""",
            (
                content.source_id,
                content.summary,
                json.dumps(content.relations),
                content.full_text,
                json.dumps(content.metadata),
            ),
        )
        self._conn.commit()
        self._cache.pop(content.source_id, None)

    def store_batch(self, contents: list[SourceContent]) -> None:
        """Store multiple sources in a single transaction."""
        self._conn.executemany(
            """INSERT OR REPLACE INTO sources (source_id, summary, relations, full_text, metadata)
               VALUES (?, ?, ?, ?, ?)""",
            [
                (
                    c.source_id,
                    c.summary,
                    json.dumps(c.relations),
                    c.full_text,
                    json.dumps(c.metadata),
                )
                for c in contents
            ],
        )
        self._conn.commit()

    def _evict_if_full(self) -> None:
        """Evict oldest entries if cache is over capacity."""
        while len(self._cache) > self._CACHE_SIZE:
            # dict preserves insertion order; pop the first key
            self._cache.pop(next(iter(self._cache)))

    def retrieve(self, source_id: str) -> SourceContent | None:
        """Retrieve content for a source.

        Args:
            source_id: The source identifier.

        Returns:
            SourceContent or None if not found.
        """
        cached = self._cache.get(source_id)
        if cached is not None:
            return cached

        cursor = self._conn.execute(
            "SELECT source_id, summary, relations, full_text, metadata FROM sources WHERE source_id = ?",
            (source_id,),
        )
        row = cursor.fetchone()
        if row is None:
            return None
        content = SourceContent._from_row(row[0], row[1], row[2], row[3], row[4])
        self._cache[source_id] = content
        self._evict_if_full()
        return content

    def retrieve_batch(self, source_ids: list[str]) -> list[SourceContent]:
        """Retrieve content for multiple sources (cache-aware)."""
        if not source_ids:
            return []

        results = []
        missing = []
        for sid in source_ids:
            cached = self._cache.get(sid)
            if cached is not None:
                results.append(cached)
            else:
                missing.append(sid)

        if missing:
            placeholders = ",".join("?" for _ in missing)
            cursor = self._conn.execute(
                f"SELECT source_id, summary, relations, full_text, metadata FROM sources WHERE source_id IN ({placeholders})",
                missing,
            )
            for row in cursor:
                content = SourceContent._from_row(row[0], row[1], row[2], row[3], row[4])
                self._cache[row[0]] = content
                results.append(content)
            self._evict_if_full()

        return results

    def remove(self, source_id: str) -> bool:
        """Remove a source from the store.

        Returns:
            True if the source existed and was removed.
        """
        cursor = self._conn.execute(
            "DELETE FROM sources WHERE source_id = ?", (source_id,)
        )
        self._conn.commit()
        self._cache.pop(source_id, None)
        return cursor.rowcount > 0

    def all_ids(self) -> list[str]:
        """Return all source IDs in the store."""
        cursor = self._conn.execute("SELECT source_id FROM sources")
        return [row[0] for row in cursor]

    def to_bytes(self) -> bytes:
        """Serialise the store to bytes for .rlat persistence.

        Uses SQLite binary backup for compact output and fast restore.
        The result starts with the SQLite file header ("SQLite format 3\\0").
        """
        import os
        import tempfile

        # Write to a temp file via backup(), then read the raw bytes
        fd, tmp_path = tempfile.mkstemp(suffix=".sqlite")
        try:
            os.close(fd)
            target = sqlite3.connect(tmp_path)
            self._conn.backup(target)
            target.close()
            with open(tmp_path, "rb") as f:
                return f.read()
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

    @classmethod
    def from_bytes(cls, data: bytes) -> SourceStore:
        """Deserialise a store from bytes.

        Auto-detects format:
        - Binary SQLite (starts with "SQLite format 3\\0") — new format
        - SQL text dump (starts with "BEGIN"/"CREATE") — legacy format
        """
        store = cls.__new__(cls)
        store.path = ":memory:"
        store._conn = sqlite3.connect(":memory:", check_same_thread=False)
        store._cache = {}

        if not data:
            store._create_tables()
            return store

        # Sniff format: SQLite binary starts with "SQLite format 3\0"
        if data[:16] == b"SQLite format 3\x00":
            # Binary format: write to temp file, backup() into :memory:
            import os
            import tempfile
            fd, tmp_path = tempfile.mkstemp(suffix=".sqlite")
            try:
                os.close(fd)
                with open(tmp_path, "wb") as f:
                    f.write(data)
                source = sqlite3.connect(tmp_path)
                source.backup(store._conn)
                source.close()
            finally:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
        else:
            # Legacy SQL text dump format
            sql = data.decode("utf-8")
            store._conn.executescript(sql)

        return store

    def close(self) -> None:
        """Close the database connection."""
        self._conn.close()

    def __del__(self) -> None:
        try:
            self._conn.close()
        except Exception:
            pass


class LosslessStore:
    """Abstract base for lossless stores that resolve source_ids to files.

    The field + registry provide semantic indexing; a LosslessStore reads
    whole files on retrieval and re-chunks them, preserving the ability to
    apply new chunkers, widen windows, or detect drift. Subclasses override
    two hooks:

      - ``_resolve_path(source_id)`` — translate a source_id into whatever
        "key" ``_read_file`` wants (a ``Path`` for local disk, a relative
        string for bundled/remote).
      - ``_read_file(key)`` — return the decoded text for that key.

    Everything else — manifest handling, chunking, drift detection, caches,
    SourceContent construction — is shared across modes so every retrieval
    feature (re-chunk, drift, widen, format dispatch) works identically in
    every backend.
    """

    def __init__(
        self,
        manifest: dict[str, Any] | None = None,
        meta_store: SourceStore | None = None,
    ) -> None:
        """
        Args:
            manifest: source_id -> relative file path (str) or rich dict
                ({source_file, heading, char_offset, content_hash, ...}).
                Stored in the knowledge model as ``__source_manifest__``.
            meta_store: Optional embedded metadata store for ``__``-prefixed
                entries (encoder config, source manifest, profile, etc.).
        """
        self._manifest: dict[str, Any] = manifest or {}
        self._meta_store: SourceStore | None = meta_store
        self._file_cache: dict[str, str] = {}
        self._chunk_cache: dict[str, list] = {}  # display-key -> list[Chunk]
        # A4: per-source drift status, populated lazily as retrieve() runs.
        # Values: "ok" (live chunk matches manifest content_hash),
        #         "drifted" (hash mismatch — file edited since build),
        #         "missing" (file not resolvable),
        #         "unknown" (no content_hash in manifest — pre-A3 build).
        self._drift_status: dict[str, str] = {}
        # Dedup warnings so users don't see one line per chunk when a single
        # file drifts — we warn once per file.
        self._drift_warned_files: set[str] = set()

    @property
    def count(self) -> int:
        return 0  # Lossless stores have no pre-materialised content count

    # ── Hooks that subclasses implement ────────────────────────────────

    def _resolve_path(self, source_id: str) -> Any | None:
        """Return the backend-specific key for ``_read_file`` (or None).

        Subclasses override. The return type is whatever that backend's
        ``_read_file`` accepts — a ``pathlib.Path`` for LocalStore, a
        relative ``str`` for BundledStore / RemoteStore.
        """
        raise NotImplementedError

    def _read_file(self, key: Any) -> str:
        """Return the decoded text for a backend-specific key.

        Subclasses override. Format dispatch (csv/pdf/docx/xlsx/plain)
        typically lives here.
        """
        raise NotImplementedError

    def _display_path(self, key: Any) -> str:
        """Human-readable rendering of a key for warnings / metadata."""
        return str(key)

    def _missing_summary(self, source_id: str) -> str:
        """Summary text used when a source_id fails to resolve."""
        return (
            f"[lossless store: file for {source_id} not resolvable. "
            f"Run `rlat refresh` to drop dangling chunks.]"
        )

    # ── Shared concrete logic ──────────────────────────────────────────

    @staticmethod
    def _parse_chunk_index(source_id: str) -> int | None:
        """Extract chunk index from source_id (last _NNNN segment)."""
        parts = source_id.rsplit("_", 1)
        if len(parts) == 2:
            try:
                return int(parts[1])
            except ValueError:
                pass
        return None

    def _manifest_entry(self, source_id: str) -> dict | None:
        """Return the manifest entry for a source_id, or None. Entries may
        be rich dicts (A2+) or bare strings (pre-A2); normalise to a dict
        shape here so callers can do uniform key lookups."""
        raw = self._manifest.get(source_id)
        if isinstance(raw, dict):
            return raw
        if isinstance(raw, str):
            return {"source_file": raw}
        return None

    def _get_chunks(self, key: Any) -> list:
        """Read and chunk the file at ``key``, caching the result."""
        cache_key = self._display_path(key)
        if cache_key not in self._chunk_cache:
            from resonance_lattice.chunker import auto_chunk
            if cache_key not in self._file_cache:
                self._file_cache[cache_key] = self._read_file(key)
            self._chunk_cache[cache_key] = auto_chunk(
                self._file_cache[cache_key], source_file=cache_key,
            )
        return self._chunk_cache[cache_key]

    def _check_drift(
        self, source_id: str, live_chunk, key: Any, emit_warning: bool = True
    ) -> str:
        """Compare a live chunk's content_hash against the manifest value.

        Returns the drift status string ("ok" / "drifted" / "unknown") and
        caches it. The first drift per file emits a one-shot RuntimeWarning
        so users notice without being spammed.
        """
        entry = self._manifest_entry(source_id)
        expected = entry.get("content_hash") if entry else None
        if not expected:
            status = "unknown"
        elif live_chunk.content_hash == expected:
            status = "ok"
        else:
            status = "drifted"
        self._drift_status[source_id] = status

        if status == "drifted" and emit_warning:
            display = self._display_path(key)
            if display not in self._drift_warned_files:
                self._drift_warned_files.add(display)
                import warnings
                # Best-effort short name for the warning — Path has .name,
                # a relative-string key just uses itself.
                short = getattr(key, "name", None) or display
                warnings.warn(
                    f"Content drift detected in {short}: chunk hashes "
                    f"no longer match the cartridge. Results may reflect "
                    f"out-of-date indexing. Run `rlat refresh` to re-sync.",
                    RuntimeWarning,
                    stacklevel=4,
                )
        return status

    def retrieve(self, source_id: str) -> SourceContent | None:
        """Retrieve content for a source by reading from the backend."""
        if source_id.startswith("__"):
            if self._meta_store is not None:
                return self._meta_store.retrieve(source_id)
            return None

        key = self._resolve_path(source_id)
        if key is None:
            self._drift_status[source_id] = "missing"
            return SourceContent(
                source_id=source_id,
                summary=self._missing_summary(source_id),
                full_text="",
                metadata={"resolved": False},
            )

        # Re-chunk the file and return the matching chunk
        chunk_idx = self._parse_chunk_index(source_id)
        chunks = self._get_chunks(key)
        display = self._display_path(key)

        if chunk_idx is not None and 0 <= chunk_idx < len(chunks):
            chunk = chunks[chunk_idx]
            drift_status = self._check_drift(source_id, chunk, key)
            return SourceContent(
                source_id=source_id,
                summary=chunk.text,
                full_text=chunk.text,
                metadata={
                    "source_file": display,
                    "heading": chunk.heading,
                    "chunk_type": chunk.chunk_type,
                    "resolved": True,
                    "drift_status": drift_status,
                },
            )

        # Fallback: file changed and chunk index is out of range
        if display not in self._file_cache:
            self._file_cache[display] = self._read_file(key)
        text = self._file_cache[display]
        return SourceContent(
            source_id=source_id,
            summary=text[:200],
            full_text=text,
            metadata={
                "source_file": display,
                "resolved": True,
            },
        )

    def retrieve_batch(self, source_ids: list[str]) -> list[SourceContent]:
        """Retrieve content for multiple sources."""
        results = []
        for sid in source_ids:
            content = self.retrieve(sid)
            if content is not None:
                results.append(content)
        return results

    def store(self, content: SourceContent) -> None:
        """Route __-prefixed entries into the metadata store; drop content.

        Chunk-text writes are ignored (the backend already owns the
        bytes), but reserved ``__``-prefixed records — encoder config,
        source manifest, remote origin, profile, retrieval config —
        need a home so that ``save()`` can serialise them with the
        knowledge model. If no meta_store is attached, this degenerates to
        the historical no-op behaviour.
        """
        if content.source_id.startswith("__") and self._meta_store is not None:
            self._meta_store.remove(content.source_id)
            self._meta_store.store(content)

    def store_batch(self, contents: list[SourceContent]) -> None:
        """Batch wrapper for :meth:`store` — routes reserved ids to meta."""
        for content in contents:
            self.store(content)

    def remove(self, source_id: str) -> bool:
        """Route reserved-id removes to the metadata store; else no-op."""
        if source_id.startswith("__") and self._meta_store is not None:
            return self._meta_store.remove(source_id)
        return False

    def all_ids(self) -> list[str]:
        """Enumerate via the manifest (source_id -> source_file map).

        The manifest is populated at load for every LosslessStore subclass
        (LocalStore, BundledStore, RemoteStore) from ``__source_manifest__``
        and is authoritative for which source_ids exist in the cartridge.
        Reserved ``__``-prefixed keys are included; callers (e.g. ``rlat
        ls``) filter them out where appropriate.
        """
        return list(self._manifest.keys())

    def to_bytes(self) -> bytes:
        """Default: no bytes to serialise. Bundled mode overrides."""
        return b""

    def drift_summary(self) -> dict[str, int]:
        """Return current drift-status counts across every source_id that
        has been retrieved (and therefore checked) so far in this session.

        Does NOT trigger retrieval. Call `verify_all()` first if you want
        a full audit. Returns counts keyed by status: ok / drifted /
        missing / unknown. Source_ids never touched don't contribute.
        """
        counts = {"ok": 0, "drifted": 0, "missing": 0, "unknown": 0}
        for status in self._drift_status.values():
            counts[status] = counts.get(status, 0) + 1
        return counts

    def verify_all(self, emit_warnings: bool = False) -> dict[str, int]:
        """Walk every real (non-reserved) manifest entry, retrieve it, and
        record drift status. Returns the same shape as drift_summary().

        `emit_warnings=False` by default — bulk verification is typically
        driven by `rlat info --verify` where we prefer to summarise in one
        print rather than stream a RuntimeWarning per file. Set True when
        the caller wants visible warnings as drifts are discovered.
        """
        for source_id in list(self._manifest.keys()):
            if source_id.startswith("__"):
                continue
            # Fast path: if we've already classified this sid, skip.
            if source_id in self._drift_status:
                continue
            # Use the normal retrieve path so _check_drift records the
            # status. Temporarily mute the per-file warning when the
            # caller asked for silence.
            if emit_warnings:
                self.retrieve(source_id)
            else:
                # Bypass the warn-once machinery for the bulk audit:
                # pre-mark the file as already-warned so _check_drift
                # records status without calling warnings.warn.
                key = self._resolve_path(source_id)
                if key is None:
                    self._drift_status[source_id] = "missing"
                    continue
                display = self._display_path(key)
                pre_warned = display in self._drift_warned_files
                self._drift_warned_files.add(display)
                try:
                    self.retrieve(source_id)
                finally:
                    if not pre_warned:
                        # Keep the silence scoped to this audit: don't
                        # prevent a future per-retrieve warning from
                        # firing on the first hot-path access.
                        self._drift_warned_files.discard(display)
        return self.drift_summary()

    def close(self) -> None:
        """No-op."""
        pass


class LocalStore(LosslessStore):
    """Lossless store backed by a local filesystem directory.

    Works with storeless .rlat knowledge models: the field + registry provide
    semantic indexing, while this store reads content from source files
    on disk at query time.

    Source IDs are mapped back to files via a manifest (source_id -> file
    path) saved in the knowledge model metadata, or by heuristic filename matching.
    """

    def __init__(
        self,
        source_root: str | Path,
        manifest: dict[str, Any] | None = None,
        meta_store: SourceStore | None = None,
    ) -> None:
        """
        Args:
            source_root: Root directory for resolving source files.
            manifest: Optional mapping of source_id -> relative file path.
                Stored in the knowledge model as __source_manifest__ when building
                with store_mode="local" (or legacy "external").
            meta_store: Optional embedded metadata store for __-prefixed
                entries (encoder config, source manifest, profile, etc.).
        """
        super().__init__(manifest=manifest, meta_store=meta_store)
        self.source_root = Path(source_root).resolve()

    def _resolve_path(self, source_id: str) -> Path | None:
        """Try to resolve a source_id to a file path under source_root."""
        # 1. Use manifest if available (fast O(1) lookup)
        if source_id in self._manifest:
            sf = self._manifest[source_id]
            # Rich manifest entries are dicts with source_file key
            if isinstance(sf, dict):
                sf = sf.get("source_file", "")
            if not sf:
                return None
            candidate = self.source_root / sf
            if candidate.exists():
                return candidate
            # Try the path as-is (absolute)
            p = Path(sf)
            if p.exists():
                return p

        # 2. Heuristic: parse source_id "{stem}_{slug}_{hash}_{idx}"
        #    Only scan direct children to avoid slow rglob on large trees.
        if not self.source_root.exists():
            return None
        parts = source_id.rsplit("_", 2)  # last 2 parts: hash, idx
        if len(parts) >= 3:
            stem_part = parts[0].split("_")[0].lower()
            for f in self.source_root.iterdir():
                if f.is_file() and f.stem.lower() == stem_part:
                    return f

        return None

    def _read_file(self, path: Path) -> str:
        """Read file content, dispatching by extension.

        Text files are read directly. Binary formats (.docx, .pdf, .xlsx)
        use optional dependencies — if missing, a helpful message is returned
        instead of raising.
        """
        ext = path.suffix.lower()

        if ext == ".csv":
            return self._read_csv(path)
        elif ext == ".tsv":
            return self._read_csv(path, delimiter="\t")
        elif ext == ".docx":
            return self._read_docx(path)
        elif ext == ".pdf":
            return self._read_pdf(path)
        elif ext in (".xlsx", ".xls"):
            return self._read_xlsx(path)
        else:
            # Default: read as UTF-8 text (covers .txt, .md, .py, etc.)
            return path.read_text(encoding="utf-8", errors="replace")

    def _missing_summary(self, source_id: str) -> str:
        # A8: actionable summary — tells the user what to do next when a
        # chunk points at a file that's no longer resolvable. `rlat refresh`
        # drops dangling chunks; `--source-root` fixes the common case of
        # a moved corpus.
        return (
            f"[external: file for {source_id} not found under "
            f"{self.source_root}. Pass --source-root <dir> if the "
            f"corpus moved, or run `rlat refresh` to drop chunks "
            f"for files that no longer exist.]"
        )

    @staticmethod
    def _read_csv(path: Path, delimiter: str = ",") -> str:
        """Read CSV/TSV as pipe-separated text for readability.

        CSVs in the wild can contain single fields well beyond Python's default
        csv.field_size_limit (131072 bytes) — e.g. the anthropic-cookbook has
        CSVs with long embedded JSON blobs. Lift the limit to sys.maxsize so
        build never aborts halfway through a corpus over one oversized field.
        """
        import sys
        try:
            csv.field_size_limit(sys.maxsize)
        except OverflowError:
            csv.field_size_limit(2**31 - 1)  # Windows 32-bit C long cap
        rows = []
        with open(path, newline="", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f, delimiter=delimiter)
            for row in reader:
                rows.append(" | ".join(row))
        return "\n".join(rows)

    @staticmethod
    def _read_docx(path: Path) -> str:
        """Read .docx using python-docx (optional dependency)."""
        try:
            from docx import Document
        except ImportError:
            return f"[cannot read {path.name}: pip install python-docx]"
        try:
            doc = Document(str(path))
            return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except Exception as e:
            return f"[error reading {path.name}: {e}]"

    @staticmethod
    def _read_pdf(path: Path) -> str:
        """Read .pdf using pymupdf or pdfplumber (optional dependency)."""
        # Try pymupdf first (fast), then pdfplumber
        try:
            import fitz  # pymupdf
            doc = fitz.open(str(path))
            pages = [page.get_text() for page in doc]
            doc.close()
            return "\n\n".join(pages)
        except ImportError:
            pass
        except Exception as e:
            return f"[error reading {path.name}: {e}]"
        try:
            import pdfplumber
            with pdfplumber.open(str(path)) as pdf:
                return "\n\n".join(
                    page.extract_text() or "" for page in pdf.pages
                )
        except ImportError:
            return f"[cannot read {path.name}: pip install pymupdf or pdfplumber]"
        except Exception as e:
            return f"[error reading {path.name}: {e}]"

    @staticmethod
    def _read_xlsx(path: Path) -> str:
        """Read .xlsx using openpyxl (optional dependency)."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            return f"[cannot read {path.name}: pip install openpyxl]"
        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)
        except Exception as e:
            return f"[error reading {path.name}: {e}]"
        sheets = []
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(" | ".join(str(c) if c is not None else "" for c in row))
            if rows:
                sheets.append(f"## {ws.title}\n" + "\n".join(rows))
        wb.close()
        return "\n\n".join(sheets)


class BundledStore(LosslessStore):
    """Lossless store backed by raw source files packed into the knowledge model.

    Replaces the deprecated embedded SQLite store while preserving every
    retrieval-quality feature: the knowledge model carries the full source
    files (zstd-compressed, individually framed) so the retriever can
    re-chunk on demand, widen windows, and run drift detection the same
    way local / remote mode do.

    Not to be confused with the legacy embedded ``SourceStore``, which
    stored pre-chunked text and lost whole-file context. BundledStore
    stores whole files — same semantics as LocalStore, same manifest
    shape, same ``_check_drift`` hash comparison — just with the bytes
    packed inside the ``.rlat`` instead of on disk.
    """

    def __init__(
        self,
        blob_reader,
        manifest: dict[str, Any] | None = None,
        meta_store: SourceStore | None = None,
    ) -> None:
        """
        Args:
            blob_reader: Instance of
                ``resonance_lattice.bundled.BlobReader`` that holds the
                index + compressed frames and decodes single files on
                demand.
            manifest: ``source_id -> rich dict`` mapping (same shape as
                LocalStore / external mode).
            meta_store: Optional embedded metadata store for
                ``__``-prefixed entries.
        """
        super().__init__(manifest=manifest, meta_store=meta_store)
        self._blob = blob_reader

    @property
    def count(self) -> int:
        return len(self._blob.paths)

    def _resolve_path(self, source_id: str) -> str | None:
        entry = self._manifest_entry(source_id)
        if entry is None:
            return None
        rel = entry.get("source_file", "")
        if not rel:
            return None
        return rel if rel in self._blob else None

    def _read_file(self, rel_path: str) -> str:
        """Decompress the blob for ``rel_path`` and decode to text.

        Text formats are decoded directly from bytes. Binary formats
        (.csv, .tsv, .docx, .pdf, .xlsx) delegate to LocalStore's static
        readers via a short-lived temp file so every existing format
        works without a second decoder implementation.
        """
        data = self._blob.read_bytes(rel_path)
        ext = Path(rel_path).suffix.lower()
        if ext in (".csv", ".tsv", ".docx", ".pdf", ".xlsx", ".xls"):
            return _decode_binary_via_tempfile(data, rel_path)
        return data.decode("utf-8", errors="replace")

    def _display_path(self, rel_path: str) -> str:
        # Bundled keys are already posix-style relative strings.
        return rel_path

    def _missing_summary(self, source_id: str) -> str:
        return (
            f"[bundled: file for {source_id} not present in the cartridge "
            f"blob store. Rebuild the cartridge with the same source tree "
            f"to include it.]"
        )

    def all_ids(self) -> list[str]:
        """Bundled mode knows exactly which files it carries — return them."""
        return list(self._blob.paths)


def _decode_binary_via_tempfile(data: bytes, rel_path: str) -> str:
    """Write ``data`` to a temp file and delegate to LocalStore's decoders.

    Used by ``BundledStore._read_file`` for binary formats whose decoders
    (python-docx, pymupdf, openpyxl) want a file path. The tempfile is
    deleted before return, so the cost is one write+read per
    binary-format passage — a fair trade to avoid duplicating every
    format decoder.
    """
    import os
    import tempfile
    suffix = Path(rel_path).suffix.lower()
    fd, tmp_path = tempfile.mkstemp(suffix=suffix or ".bin")
    try:
        with os.fdopen(fd, "wb") as f:
            f.write(data)
        tmp = Path(tmp_path)
        if suffix == ".csv":
            return LocalStore._read_csv(tmp)
        if suffix == ".tsv":
            return LocalStore._read_csv(tmp, delimiter="\t")
        if suffix == ".docx":
            return LocalStore._read_docx(tmp)
        if suffix == ".pdf":
            return LocalStore._read_pdf(tmp)
        if suffix in (".xlsx", ".xls"):
            return LocalStore._read_xlsx(tmp)
        return data.decode("utf-8", errors="replace")
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


class RemoteStore(LosslessStore):
    """Lossless store backed by a public HTTP origin (GitHub) with cache.

    The knowledge model pins to a commit SHA at build time. Query-path reads
    go through a two-tier cache:

        _file_cache (in-memory LRU, inherited from LosslessStore)
          → DiskCache (persistent, ~/.cache/rlat/remote/<origin>/<sha>/...)
            → fetcher.fetch(sha, rel_path) (network, only on miss)

    SHA-pinning is what makes the disk cache safe: once bytes for a
    given (origin, sha, rel_path) tuple land on disk, they are
    immutable, so we skip revalidation entirely and warm queries feel
    like local mode.

    Load/query paths NEVER call the network beyond the cache. Freshness
    is strictly opt-in via ``rlat freshness`` (drift check) and
    ``rlat sync`` (explicit upgrade) — the lockfile model from the
    three-mode-store plan.
    """

    def __init__(
        self,
        fetcher,
        commit_sha: str,
        cache,
        origin_key: str,
        manifest: dict[str, Any] | None = None,
        meta_store: SourceStore | None = None,
    ) -> None:
        """
        Args:
            fetcher: Implementation of ``resonance_lattice.remote.Fetcher``
                — currently ``GithubFetcher``.
            commit_sha: 40-char commit SHA the knowledge model pins to.
            cache: ``resonance_lattice.remote.DiskCache`` instance.
            origin_key: Filesystem-safe identifier for the origin (used
                as a disk-cache namespace). Typically
                ``f"github__{org}__{repo}"`` — see
                ``GithubOrigin.key``.
            manifest: source_id → rich dict mapping, same shape as
                LocalStore / BundledStore.
            meta_store: Optional embedded metadata SourceStore for
                ``__``-prefixed entries (encoder config,
                ``__remote_origin__``, profile, ...).
        """
        super().__init__(manifest=manifest, meta_store=meta_store)
        self._fetcher = fetcher
        self._sha = commit_sha
        self._cache = cache
        self._origin_key = origin_key

    @property
    def commit_sha(self) -> str:
        return self._sha

    @property
    def origin_key(self) -> str:
        return self._origin_key

    def _resolve_path(self, source_id: str) -> str | None:
        entry = self._manifest_entry(source_id)
        if entry is None:
            return None
        rel = entry.get("source_file", "")
        return rel or None

    def _read_file(self, rel_path: str) -> str:
        """Return decoded text for ``rel_path``. Cache-first, network on miss.

        Errors on the network fall through to a human-readable placeholder
        so a single flaky fetch doesn't crash the whole query; the
        surrounding retrieve() marks the source as resolved=False and
        downstream callers render the placeholder.
        """
        # In-memory LRU is handled by LosslessStore._get_chunks, which
        # keys on display_path (rel_path). Here we only consult the
        # persistent disk cache + fetch on miss.
        cached = self._cache.get(self._origin_key, self._sha, rel_path)
        if cached is not None:
            data = cached
        else:
            try:
                data = self._fetcher.fetch(self._sha, rel_path)
            except Exception as e:
                import warnings
                warnings.warn(
                    f"Remote fetch failed for {rel_path} at {self._sha[:8]} "
                    f"({type(e).__name__}: {e}). Returning empty passage; "
                    f"retry after the network recovers.",
                    RuntimeWarning,
                    stacklevel=4,
                )
                return ""
            try:
                self._cache.put(self._origin_key, self._sha, rel_path, data)
            except OSError:
                # Cache write failed (disk full / permission); still
                # return the fetched bytes for this query.
                pass
        ext = Path(rel_path).suffix.lower()
        if ext in (".csv", ".tsv", ".docx", ".pdf", ".xlsx", ".xls"):
            return _decode_binary_via_tempfile(data, rel_path)
        return data.decode("utf-8", errors="replace")

    def _display_path(self, rel_path: str) -> str:
        return rel_path

    def _missing_summary(self, source_id: str) -> str:
        return (
            f"[remote: file for {source_id} not found in the cartridge "
            f"manifest at pinned sha {self._sha[:8]}. Run `rlat sync` "
            f"to update the cartridge to upstream HEAD.]"
        )


# Backward-compat alias. ``ExternalStore`` was the name before the three-mode
# refactor (feat/three-mode-store). Call-site renames happen in Phase 6
# alongside the user-facing ``--store-mode`` rename (external → local). Until
# then, every `ExternalStore` import keeps working unchanged.
ExternalStore = LocalStore
